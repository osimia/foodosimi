# fruitsite/orders/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login
from django.contrib import messages
from django.utils.crypto import get_random_string
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
import io
import json
from datetime import datetime
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from cart.models import CartItem
from cart.views import get_cart_owner_info
from accounts.models import User
from .models import Order, OrderItem

def checkout(request):
    # Получаем корзину на основе авторизации или сессии
    owner_filter, _ = get_cart_owner_info(request)
    cart_items = CartItem.objects.filter(**owner_filter)
    
    if not cart_items.exists():
        messages.info(request, "Ваша корзина пуста.")
        return redirect('products:product_list')

    if request.method == 'POST':
        phone = request.POST.get('phone')
        delivery_address = request.POST.get('delivery_address')
        
        # Базовая валидация
        if not phone or not delivery_address:
            messages.error(request, "Пожалуйста, заполните все обязательные поля.")
            return redirect('orders:checkout')
            
        # Очистка номера телефона от нецифровых символов
        clean_phone = re.sub(r'\D', '', phone)
        if len(clean_phone) < 10:
            messages.error(request, "Пожалуйста, введите корректный номер телефона (минимум 10 цифр).")
            return redirect('orders:checkout')
            
        user = request.user
        
        # Если пользователь не авторизован, создаем новый аккаунт или находим существующий
        if not request.user.is_authenticated:
            # Ищем пользователя с таким телефоном
            existing_user = User.objects.filter(phone=clean_phone).first()
            
            if existing_user:
                user = existing_user
                # Если нашли пользователя, авторизуем его
                login(request, user)
                messages.success(request, f"Вход выполнен по номеру телефона {phone}.")
            else:
                # Создаем нового пользователя
                username = f"user_{get_random_string(8)}"
                password = get_random_string(12)
                user = User.objects.create_user(
                    username=username,
                    password=password,
                    phone=clean_phone,
                    role='buyer'
                )
                login(request, user)
                messages.success(request, f"Аккаунт создан автоматически. Ваш логин: {username}. Пароль отправлен на ваш номер телефона.")
                
            # Перенос товаров из анонимной корзины в корзину пользователя
            if 'session_key' in owner_filter:
                session_cart_items = CartItem.objects.filter(session_key=request.session.session_key)
                for item in session_cart_items:
                    item.buyer = user
                    item.session_key = None
                    item.save()
        
        total_amount = sum(item.product.price * item.quantity for item in cart_items)

        # Создаем заказ
        order = Order.objects.create(
            buyer=user,
            phone_number=clean_phone,
            delivery_address=delivery_address,
            total_amount=total_amount
        )

        # Добавляем товары в заказ
        for item in cart_items:
            OrderItem.objects.create(
                order=order,
                product=item.product,
                quantity=item.quantity,
                price=item.product.price
            )
            
        # Очищаем корзину
        cart_items.delete()
        
        # Разные сообщения для новых и существующих пользователей
        if not request.user.is_authenticated or not hasattr(request.user, 'last_login') or request.user.last_login is None:
            messages.success(request, f"Заказ #{order.id} успешно оформлен! Добро пожаловать в личный кабинет, где вы можете отслеживать статус своих заказов.")
        else:
            messages.success(request, f"Заказ #{order.id} успешно оформлен! Вы можете отслеживать статус заказа в своем профиле.")
        
        return redirect('accounts:profile')

    # Подсчитываем общую сумму
    total = sum(item.product.price * item.quantity for item in cart_items)
    
    return render(request, 'orders/checkout.html', {
        'cart_items': cart_items,
        'total': total
    })


def my_orders(request):
    # Если пользователь авторизован, показываем его заказы
    if request.user.is_authenticated:
        orders = Order.objects.filter(buyer=request.user).order_by('-created_at')
        return render(request, 'orders/my_orders.html', {'orders': orders})
    
    # Для неавторизованных пользователей
    if request.method == 'POST':
        phone = request.POST.get('phone')
        if phone:
            # Очистка номера телефона от нецифровых символов
            clean_phone = re.sub(r'\D', '', phone)
            if len(clean_phone) >= 10:
                orders = Order.objects.filter(phone_number=clean_phone).order_by('-created_at')
                if orders:
                    return render(request, 'orders/my_orders.html', {
                        'orders': orders,
                        'guest_phone': phone
                    })
                else:
                    messages.error(request, "По указанному номеру телефона заказов не найдено.")
            else:
                messages.error(request, "Пожалуйста, введите корректный номер телефона.")
    
    # По умолчанию показываем форму для ввода телефона
    return render(request, 'orders/find_orders.html')


def find_orders(request):
    """
    Отдельная страница для поиска заказов по номеру телефона.
    Позволяет неавторизованным пользователям найти свои заказы.
    """
    orders = None
    phone = None
    
    if request.method == 'POST':
        phone = request.POST.get('phone')
        if phone:
            # Очистка номера телефона от нецифровых символов
            clean_phone = re.sub(r'\D', '', phone)
            if len(clean_phone) >= 10:
                orders = Order.objects.filter(phone_number=clean_phone).order_by('-created_at')
                if not orders:
                    messages.error(request, "По указанному номеру телефона заказов не найдено.")
            else:
                messages.error(request, "Пожалуйста, введите корректный номер телефона.")
    
    return render(request, 'orders/find_orders.html', {
        'orders': orders,
        'phone': phone
    })


@login_required
def seller_orders(request):
    """
    Представление для продавцов - показывает заказы на их товары
    """
    if request.user.role != 'master':
        messages.error(request, "У вас нет доступа к этой странице.")
        return redirect('accounts:my_products')
    
    # Получаем все заказы, которые содержат товары этого продавца
    order_items = OrderItem.objects.filter(
        product__master=request.user
    ).select_related('order', 'product').order_by('-order__created_at')
    
    # Группируем по заказам для удобного отображения
    orders_dict = {}
    for item in order_items:
        order_id = item.order.id
        if order_id not in orders_dict:
            orders_dict[order_id] = {
                'order': item.order,
                'items': []
            }
        orders_dict[order_id]['items'].append(item)
    
    # Преобразуем в список для шаблона
    orders_data = list(orders_dict.values())
    
    return render(request, 'orders/seller_orders.html', {
        'orders_data': orders_data
    })


@login_required
def accept_order(request, order_id):
    """
    Принятие заказа продавцом
    """
    if request.user.role != 'master':
        messages.error(request, "У вас нет доступа к этой функции.")
        return redirect('orders:seller_orders')
    
    order = get_object_or_404(Order, id=order_id)
    
    # Проверяем, что в заказе есть товары этого продавца
    seller_items = OrderItem.objects.filter(
        order=order,
        product__master=request.user
    )
    
    if not seller_items.exists():
        messages.error(request, "У вас нет товаров в этом заказе.")
        return redirect('orders:seller_orders')
    
    if order.status == 'new':
        order.status = 'accepted'
        order.save()
        messages.success(request, f"Заказ #{order.id} успешно принят.")
    else:
        messages.info(request, f"Заказ #{order.id} уже обработан.")
    
    return redirect('orders:seller_orders')

@login_required
def reject_order(request, order_id):
    """
    Отклонение заказа продавцом
    """
    if request.user.role != 'master':
        messages.error(request, "У вас нет доступа к этой функции.")
        return redirect('orders:seller_orders')
    
    order = get_object_or_404(Order, id=order_id)
    
    # Проверяем, что в заказе есть товары этого продавца
    seller_items = OrderItem.objects.filter(
        order=order,
        product__master=request.user
    )
    
    if not seller_items.exists():
        messages.error(request, "У вас нет товаров в этом заказе.")
        return redirect('orders:seller_orders')
    
    if order.status == 'new':
        order.status = 'rejected'
        order.save()
        messages.success(request, f"Заказ #{order.id} отклонен.")
    else:
        messages.info(request, f"Заказ #{order.id} уже обработан.")
    
    return redirect('orders:seller_orders')

@login_required
def update_order_status(request, order_id):
    """
    Обновление статуса заказа (для перевода в обработку, отправку и т.д.)
    """
    if request.user.role != 'master':
        messages.error(request, "У вас нет доступа к этой функции.")
        return redirect('orders:seller_orders')
    
    order = get_object_or_404(Order, id=order_id)
    
    # Проверяем, что в заказе есть товары этого продавца
    seller_items = OrderItem.objects.filter(
        order=order,
        product__master=request.user
    )
    
    if not seller_items.exists():
        messages.error(request, "У вас нет товаров в этом заказе.")
        return redirect('orders:seller_orders')
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        tracking_number = request.POST.get('tracking_number', '')
        
        # Проверяем допустимые статусы для изменения
        allowed_statuses = ['processing', 'shipped', 'delivered']
        if new_status in allowed_statuses:
            order.status = new_status
            if tracking_number:
                order.tracking_number = tracking_number
            order.save()
            messages.success(request, f"Статус заказа #{order.id} обновлен.")
        else:
            messages.error(request, "Недопустимый статус заказа.")
    
    return redirect('orders:seller_orders')

@login_required
def confirm_delivery(request, order_id):
    """
    Подтверждение получения заказа покупателем
    """
    order = get_object_or_404(Order, id=order_id)
    
    # Проверяем, что это покупатель этого заказа
    if order.buyer != request.user:
        messages.error(request, "У вас нет доступа к этому заказу.")
        return redirect('orders:my_orders')
    
    # Проверяем, что заказ в статусе "Отправлен"
    if order.status != 'shipped':
        messages.error(request, "Этот заказ нельзя подтвердить как полученный.")
        return redirect('orders:my_orders')
    
    if request.method == 'POST':
        # Меняем статус на "Доставлен"
        order.status = 'delivered'
        order.save()
        messages.success(request, f"Спасибо! Заказ #{order.id} отмечен как доставленный.")
    
    return redirect('orders:my_orders')

@login_required
def download_invoice(request, order_id):
    """
    Скачивание накладной заказа в формате Excel
    """
    if not OPENPYXL_AVAILABLE:
        messages.error(request, "Функция скачивания накладных временно недоступна.")
        return redirect('orders:seller_orders' if request.user.role == 'master' else 'orders:my_orders')
    
    order = get_object_or_404(Order, id=order_id)
    
    # Проверяем права доступа
    if request.user.role == 'master':
        # Для мастера - проверяем что в заказе есть его товары
        seller_items = OrderItem.objects.filter(
            order=order,
            product__master=request.user
        )
        if not seller_items.exists():
            messages.error(request, "У вас нет доступа к этому заказу.")
            return redirect('orders:seller_orders')
    else:
        # Для покупателя - проверяем что это его заказ
        if order.buyer != request.user:
            messages.error(request, "У вас нет доступа к этому заказу.")
            return redirect('orders:my_orders')

    # Создаем Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = f"Накладная_{order.id}"

    # Стили
    title_font = Font(name='Arial', size=16, bold=True)
    header_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=11)
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    # Границы
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовок документа
    ws.merge_cells('A1:G1')
    ws['A1'] = 'ТОВАРНАЯ НАКЛАДНАЯ'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment
    
    # Номер и дата накладной
    current_row = 3
    ws[f'A{current_row}'] = f'Накладная № {order.id}'
    ws[f'A{current_row}'].font = header_font
    
    ws[f'E{current_row}'] = f'от {order.created_at.strftime("%d.%m.%Y")}'
